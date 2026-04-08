"""
아임웹 주문/결제 백업을 Payload Orders 컬렉션으로 일괄 등록.

- 주문서.xlsx: 주문 + 품목 (한 주문에 여러 행 가능)
- 결제내역.xlsx: 주문번호 ↔ 결제수단/상태 매핑

처리 로직:
1. 결제내역 → 주문번호 키로 dict
2. 주문서를 주문번호로 그룹핑
3. 그룹마다 첫 행 정보 + 품목명 합쳐서 Order 1건 생성
4. buyerEmail로 user 찾아서 relationship 연결 (없으면 user=null)
5. Payload /api/orders POST
"""
import openpyxl
import requests
import json
import sys
import os
from datetime import datetime

# 설정
ORDERS_XLSX = r"C:\Users\USER\Desktop\AI놀자_백업\기본_양식_20260408120238.xlsx"
PAYMENTS_XLSX = r"C:\Users\USER\Desktop\AI놀자_백업\결제_내역_20260408120233.xlsx"
API_BASE = "https://ainolza.kr"
ADMIN_EMAIL = "rex39@naver.com"
ADMIN_PASSWORD = "rudghks39"

# 상태 매핑 (아임웹 → 우리 스키마)
STATUS_MAP = {
    "거래개시": "pending",
    "결제완료": "paid",
    "이용중": "active",
    "완료": "completed",
    "환불요청": "refund_requested",
    "환불완료": "refunded",
    "결제실패": "failed",
    "주문취소": "cancelled",
    "취소": "cancelled",
    "환불": "refunded",
}

# 결제수단 매핑
PAY_METHOD_MAP = {
    "무통장입금": "vbank",
    "가상계좌": "vbank",
    "신용카드": "card",
    "카드": "card",
    "체크카드": "card",
    "카카오페이": "kakaopay",
    "네이버페이": "naverpay",
    "토스페이": "tosspay",
    "계좌이체": "trans",
    "휴대폰": "phone",
    "휴대폰결제": "phone",
}


def map_status(raw):
    if not raw:
        return "pending"
    raw = str(raw).strip()
    return STATUS_MAP.get(raw, "pending")


def map_pay_method(raw):
    if not raw:
        return None
    raw = str(raw).strip()
    return PAY_METHOD_MAP.get(raw)


def cell(row, idx_map, key):
    """row에서 안전하게 컬럼 값 가져오기."""
    if key not in idx_map:
        return None
    i = idx_map[key]
    if i is None or i >= len(row):
        return None
    return row[i]


def parse_date(value):
    """엑셀의 날짜 문자열을 ISO 형식으로 변환."""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S").isoformat()
    except Exception:
        return None


def main():
    # 1. 관리자 로그인
    print("🔐 관리자 로그인 중...")
    session = requests.Session()
    login_res = session.post(
        f"{API_BASE}/api/users/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        timeout=15,
    )
    if login_res.status_code != 200:
        print(f"❌ 로그인 실패: {login_res.status_code} {login_res.text[:200]}")
        sys.exit(1)
    print("✅ 로그인 성공")

    # 1-1. 회원 전체 가져와서 email → id 매핑
    print("👥 회원 목록 가져오는 중...")
    user_map = {}
    page = 1
    while True:
        res = session.get(f"{API_BASE}/api/users", params={"limit": 200, "page": page, "depth": 0}, timeout=20)
        if res.status_code != 200:
            print(f"❌ 회원 조회 실패: {res.status_code} {res.text[:200]}")
            sys.exit(1)
        data = res.json()
        for u in data.get("docs", []):
            email = (u.get("email") or "").strip().lower()
            if email:
                user_map[email] = u.get("id")
        if page >= data.get("totalPages", 1):
            break
        page += 1
    print(f"   → {len(user_map)}명")

    # 2. 결제내역 → orderNumber 키 dict
    print(f"📖 결제내역 읽는 중: {PAYMENTS_XLSX}")
    pay_wb = openpyxl.load_workbook(PAYMENTS_XLSX, read_only=True)
    pay_ws = pay_wb.active
    pay_rows = list(pay_ws.iter_rows(values_only=True))
    pay_headers = list(pay_rows[0])
    pay_idx = {h: i for i, h in enumerate(pay_headers)}

    payments = {}  # orderNumber -> dict
    for row in pay_rows[1:]:
        if not row or not cell(row, pay_idx, "주문번호"):
            continue
        order_no = str(cell(row, pay_idx, "주문번호")).strip()
        if order_no not in payments:
            payments[order_no] = {
                "method": cell(row, pay_idx, "결제수단"),
                "paymentStatus": cell(row, pay_idx, "결제상태"),
                "vbankName": cell(row, pay_idx, "은행명"),
                "vbankNum": cell(row, pay_idx, "계좌번호"),
                "vbankDate": cell(row, pay_idx, "만료일시"),
            }
    print(f"   → {len(payments)}개 주문에 대한 결제정보")

    # 3. 주문서 → orderNumber로 그룹핑
    print(f"📖 주문서 읽는 중: {ORDERS_XLSX}")
    ord_wb = openpyxl.load_workbook(ORDERS_XLSX, read_only=True)
    ord_ws = ord_wb.active
    ord_rows = list(ord_ws.iter_rows(values_only=True))
    ord_headers = list(ord_rows[0])
    ord_idx = {h: i for i, h in enumerate(ord_headers)}

    grouped = {}  # orderNumber -> list of rows
    for row in ord_rows[1:]:
        if not row or not cell(row, ord_idx, "주문번호"):
            continue
        order_no = str(cell(row, ord_idx, "주문번호")).strip()
        grouped.setdefault(order_no, []).append(row)

    print(f"   → {len(grouped)}개 고유 주문 (총 {len(ord_rows) - 1}개 품목 행)")

    # 4. 각 주문을 Payload에 등록
    success = 0
    skipped = 0
    failed = 0
    failed_list = []

    # CLI: --limit N (테스트용)
    limit = None
    if "--limit" in sys.argv:
        idx = sys.argv.index("--limit")
        limit = int(sys.argv[idx + 1])
        print(f"⚠️  TEST MODE: 최대 {limit}건만 처리")

    for i, (order_no, items) in enumerate(grouped.items(), 1):
        if limit and i > limit:
            break
        first = items[0]

        # 품목명 합치기
        product_names = []
        for item in items:
            name = cell(item, ord_idx, "상품명")
            opt = cell(item, ord_idx, "옵션명")
            if name:
                if opt:
                    product_names.append(f"{name} ({opt})")
                else:
                    product_names.append(str(name))
        product_name = " + ".join(product_names) if product_names else "상품명 없음"

        buyer_email = cell(first, ord_idx, "주문자 이메일")
        buyer_name = cell(first, ord_idx, "주문자 이름")
        buyer_phone = cell(first, ord_idx, "주문자 번호")
        amount = cell(first, ord_idx, "최종주문금액") or 0
        original_amount = cell(first, ord_idx, "총 품목합계금액")
        order_status_raw = cell(first, ord_idx, "주문상태")
        ordered_at = cell(first, ord_idx, "주문일")

        if not buyer_email or "@" not in str(buyer_email):
            skipped += 1
            continue

        pay = payments.get(order_no, {})
        pay_method = map_pay_method(pay.get("method"))

        # 결제내역의 결제상태 반영
        pay_status = pay.get("paymentStatus")
        if pay_status:
            mapped_pay_status = map_status(pay_status)
            # 결제완료라면 paid, 그 외엔 주문상태 우선
            if mapped_pay_status in ("paid", "refunded", "cancelled", "failed"):
                final_status = mapped_pay_status
            else:
                final_status = map_status(order_status_raw)
        else:
            final_status = map_status(order_status_raw)

        # 회원 매칭
        email_lower = str(buyer_email).strip().lower()
        user_id = user_map.get(email_lower)

        data = {
            "orderNumber": order_no,
            "buyerName": str(buyer_name) if buyer_name else "이름없음",
            "buyerEmail": str(buyer_email).strip(),
            "buyerPhone": str(buyer_phone) if buyer_phone else None,
            "productName": product_name,
            "productType": "class",
            "amount": int(amount) if amount else 0,
            "originalAmount": int(original_amount) if original_amount else None,
            "payMethod": pay_method,
            "status": final_status,
            "vbankName": str(pay.get("vbankName")) if pay.get("vbankName") else None,
            "vbankNum": str(pay.get("vbankNum")) if pay.get("vbankNum") else None,
        }
        if user_id:
            data["user"] = user_id

        # createdAt은 Payload가 자동 설정. 주문일은 adminMemo에 기록.
        if ordered_at:
            data["adminMemo"] = f"아임웹 주문일: {ordered_at}"

        # None 값 제거 (Payload validation 회피)
        data = {k: v for k, v in data.items() if v is not None}

        try:
            res = session.post(f"{API_BASE}/api/orders", json=data, timeout=15)
            if res.status_code in (200, 201):
                success += 1
                if i % 50 == 0:
                    print(f"  ... {i}/{len(grouped)} 진행 중 (성공: {success})")
            elif res.status_code == 400 and ("unique" in res.text.lower() or "duplicate" in res.text.lower()):
                skipped += 1
            else:
                failed += 1
                failed_list.append({"order": order_no, "status": res.status_code, "error": res.text[:300]})
        except Exception as e:
            failed += 1
            failed_list.append({"order": order_no, "error": str(e)})

    # 5. 결과 보고
    print()
    print("=" * 50)
    print("📋 주문 등록 결과")
    print("=" * 50)
    print(f"  ✅ 성공:   {success}건")
    print(f"  ⏭️  건너뜀: {skipped}건 (이메일 누락 또는 중복)")
    print(f"  ❌ 실패:   {failed}건")
    print(f"  📊 총합:   {success + skipped + failed}건 / 그룹 {len(grouped)}개")

    if failed_list:
        report_path = os.path.join(os.path.dirname(ORDERS_XLSX), "orders_import_failed.json")
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(failed_list, f, ensure_ascii=False, indent=2)
        print(f"\n💾 실패 목록 저장: {report_path}")


if __name__ == "__main__":
    main()
